#!/usr/bin/env python3
# coding=utf-8
"""..."""
import csv
import sys

from openpyxl import load_workbook

# https://www.degruyterbrill.com/publishing/for-librarians/ejournals

def read(filename):
    wb = load_workbook(filename = filename)
    header = None
    for i, row in enumerate(wb.worksheets[0].iter_rows()):
        if i == 3:
            header = [c.value for c in row]
        elif i > 3:
            yield dict(zip(header, [c.value for c in row]))
            #d['URL'] = row[1].hyperlink.target


def format_price(p):
    if p is None:
        return "€0", ""
    if len(p) == 0:
        return "€0", ""
    if p == 'none':
        return "€0", ""
    try:
        return "€%d" % int(p), ""
    except:
        raise ValueError("?? %s" % p)


writer = csv.writer(sys.stdout)
writer.writerow(['Date', 'Journal', 'Publisher', 'Cost', 'URL', 'Comment'])

special_cases = []
for o in read('DeGruyter_Journal_Price_List_2026__EUR__2026-02-20.xlsx'):
    try:
        price, comment = format_price(o['APC EUR'])
    except Exception as e:
        special_cases.append(o)
        continue
    writer.writerow([
        '2026-02-20',
        o['Title'].strip(),
        'De Gruyter',
        price,
        o['URL'],
        comment,
    ])

print('\n\n')
for o in special_cases:
    print(f"UNPARSED: {o['Title'].strip()} — {o['APC EUR']}", file=sys.stderr)
