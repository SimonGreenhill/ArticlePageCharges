#!/usr/bin/env python3
# coding=utf-8
"""..."""
import csv
import sys

from openpyxl import load_workbook


# from https://www.elsevier.com/open-access

def read(filename):
    wb = load_workbook(filename = filename)
    header = None
    for i, row in enumerate(wb.worksheets[0].iter_rows()):
        if row[0].value == 'ISSN':
            header = [r.value for r in row]
        elif header:
            d = dict(zip(header, [c.value for c in row]))
            yield d


writer = csv.writer(sys.stdout)
writer.writerow(['Date', 'Journal', 'Publisher', 'Cost', 'URL', 'Comment'])

for o in read('elsevier-2026-05-06.xlsx'):
    if o['USD'] == '**':
        # APC is waived (new journal), or sponsored by a third party, or invoiced directly to authors.
        # ignore these for now
        continue

    if o['USD'] is not None:  # could be 0
        amt = "$%d" % o['USD']
    elif o['EUR'] is not None:
        amt = "€%d" % o['EUR']
    elif o['GBP'] is not None:
        amt = "£%d" % o['GBP']
    else:
        raise ValueError("No price - use JPY?")  # don't want to add another currency if I don't have to

    writer.writerow([
        '2026-05-06',
        o['Title'].replace(" Article Publishing Charge", ""),
        'Elsevier',
        amt,
        f"https://www.elsevier.com/locate/issn/{o['ISSN']}",
        o['Business model'],
    ])
