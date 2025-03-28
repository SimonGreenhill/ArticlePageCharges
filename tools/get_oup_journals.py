#!/usr/bin/env python3
# coding=utf-8
"""..."""

import requests
from io import BytesIO

from openpyxl import load_workbook

URL = "https://fdslive.oup.com/www.oup.com/academic/xls/openaccess/charges.xlsx"


def load(url=URL):
    return load_workbook(filename = BytesIO(requests.get(URL).content))

def read(wb):
    header = None
    for row in wb.worksheets[0].iter_rows():
        if row[1].value == 'Journal':
            header = [c.value for c in row]
        elif header:
            yield(dict(zip(header, [c.value for c in row])))
        else:
            continue


def format_cost(o):
    curr = {'EUR': '€', 'GBP': '£', 'USD': '$', 'JPY': '¥'}.get(o['APC Base Currency'].strip())
    assert curr is not None, repr(o['APC Base Currency'])
    return "%s%s" % (curr, o["2024 price"])


def format_journal(j):
    if ',' in j:
        return '"%s"' % j
    else:
        return j


for o in read(load(URL)):
    print("2025-03-28,%s,%s,%s,%s,Oxford University Press" % (
        format_journal(o['Journal']),
        format_cost(o),
        "https://academic.oup.com/%s" % o['code'].lower() if o['code'] else '',
        ""
    ))

